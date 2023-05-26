import os
import re
from datetime import datetime
from datetime import date
from random import randint
from time import sleep as pause
import csv
from typing import Any, Iterable, Mapping, Optional

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from tqdm import tqdm
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import string

start_time = datetime.now()
errors = []
total_pagen = []
with open('страницы для сбора наличия', 'r') as file:

    urls = [x for x in file]
    x = [word.strip(string.punctuation) for word in str(urls).split()]
    for i in x:
        total_pagen.append(i)
print(total_pagen)


def get_date_and_time() -> str:
    return datetime.now().strftime('%d.%m.%y %H-%M-%S')


def to_excel(data: Iterable[Mapping[str, Any]], column_names: Iterable[str],
             file_name: Optional[str] = "table") -> None:
    """
    Создаёт из итерируемого объекта и имён столбцов
    xlsx файл в папке "resulting files".
    """
    wb = Workbook()
    worksheet = wb.active
    side = Side(border_style='thin')
    border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side
    )
    alignment = Alignment(
        horizontal='left',
        vertical='center'
    )
    column_widths = []

    for column, name in enumerate(column_names, 1):
        cell = worksheet.cell(
            column=column,
            row=1,
            value=name
        )
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill('solid', fgColor='C5D9F1')
        cell.border = border
        cell.alignment = alignment
        column_widths.append(len(name) + 2)

    for row, product in enumerate(data, 2):
        if not product:
            print(row)
            continue
        for column, name in enumerate(column_names, 1):
            cell = worksheet.cell(
                column=column,
                row=row,
                value=product.get(name, '')
            )
            cell.font = Font(name='Calibri', size=11, bold=False)
            cell.border = border
            cell.alignment = alignment
            column_widths[column -
                          1] = max(column_widths[column -
                                                 1], len(str(cell.value)))

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(
            i)].width = column_width + 1

    datetime_now = get_date_and_time()
    wb.save(f"Сбор наличия/{file_name}____{datetime_now}.xlsx")

options = webdriver.ChromeOptions()
options.page_load_strategy = 'none'
count = 1
with webdriver.Chrome(options =options) as driver:
    driver.maximize_window()

    print("Найдено страниц в каталоге:", len(total_pagen))

    pause(1)

    products = []
    column_names = [
        'артикул',
        'наименование',
        'цена',
        'наличие'
    ]


    #обработка каталогов на покарточную пагинацию

    for url in total_pagen:
        total_cards = []
        print(total_cards)
        try:

            driver.get(url)
            print('*******' * 20)
            pause(1)
            try:
                print('13131312')
                window_handles = driver.window_handles
                print(window_handles)
                # Переключаемся на второе окно
                driver.switch_to.window(window_handles[0])
                pause(2)
                element_present = EC.presence_of_element_located((By.CLASS_NAME, 'products__info-block'))
                WebDriverWait(driver, 15).until(element_present)
                soup = BeautifulSoup(driver.page_source, 'html.parser')
                cards_url = ['https://www.ORG2' + x.find('a').get('href') for x in soup.find_all('div', class_='products__info-block')]
                print(url, 'вонючее окно обошел', cards_url, 'собранные таблички11111')
                if cards_url:
                    total_cards.append(cards_url)
                    pause(2)
                else:
                    print('ALARM!!!' * 20)
                    driver.get(url)
                    window_handles = driver.window_handles
                    print(window_handles)
                    driver.switch_to.window(window_handles[0])
                    soup = BeautifulSoup(driver.page_source, 'html.parser')
                    cards_url = ['https://www.ORG2' + x.find('a').get('href') for x in soup.find_all('div', class_='products__info-block')]
                    print(url, 'вонючее окно обошел', cards_url, 'собранные таблички11111')
            except Exception:
                driver.get(url)
                element_present = EC.presence_of_element_located((By.CLASS_NAME, 'products__info-block'))
                WebDriverWait(driver, 15).until(element_present)

                soup = BeautifulSoup(driver.page_source, 'html.parser')
                pause(2)
                cards_url = ['https://www.ORG2' + x.find('a').get('href') for x in soup.find_all('div', class_='products__info-block')]
                print('except', cards_url, 'собранные таблички')
                total_cards.append(cards_url)
                pause(2)
                print('Timed out waiting for page to load')
                continue


        #работа с карточками, извлечение данных
            for card in total_cards:
                print(card, 'card')
                for url in card:
                    print(url, 'url')

                    try:
                        driver.get(url)
                        pause(randint(4,7))
                        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'js-price-inner')))
                        driver.execute_script("window.scrollBy(0,5000)")
                        soup1 = BeautifulSoup(driver.page_source, 'lxml')

                        print(url)
                        url_len = url.split('/')

                        print(len(url_len))
                        if len(url_len)==8:

                            for x in soup1.find_all('div', class_='var-header col-cont'):

                                product = dict()
                                try:
                                    article = x.find('span', class_='variant-list__jde-code variant-list__jde-num').text

                                    name = x.find('span', class_='variant-list__info-title js-variant_title').text.split('\n')[1]

                                    if any(p['артикул'] == article or p['наименование'] == name for p in products):
                                        print('такое уже записано', article,' + '*10,  name, url)
                                        continue
                                    else:
                                        product['артикул'] = article
                                        product['наименование'] = name

                                # except Exception:
                                #     product['артикул'] = 'нет данных'
                                #     product['наименование'] = 'нет данных'


                                        try:
                                            product_price = x.find('span', class_='js-price-inner').text.split('\n')[1]
                                            product_price = product_price.replace(',', '.').replace('Р','').replace(' ','')
                                            if '.' not in product_price:
                                                product_price += '.00'
                                            match = re.search(r'\d+\.\d+', product_price)
                                            if match:
                                                product['цена'] = float(match.group(0))
                                            else:
                                                product['цена'] = None
                                        except Exception:
                                            product['цена'] = None

                                        try:
                                            availability = x.find('a', class_='js-territories-popup').text
                                            availability = availability.replace(',', '.').replace('шт.','').replace(' ','')
                                            if isinstance(availability, str) and availability.isdigit():
                                                product['наличие'] = int(availability)
                                            else:
                                                raise ValueError("Некорректное значение наличия")
                                        except Exception:
                                            try:

                                                availability_tag = x.select_one('.variant-list__price-availability.available, .variant-list__price-availability.on-order-available')
                                                if availability_tag:
                                                    availability = availability_tag.get('data-variant')
                                                    product['наличие'] = availability
                                                else:
                                                    product['наличие'] = 'нет данных'
                                            except Exception:
                                                product['наличие'] = 'нет данных'

                                        count += 1

                                        products.append(product)
                                        print(product)
                                        print(f'{count}    КАРТОЧКА ОБРАБОТАНА')


                                        print('+++++++++++')
                                except Exception:
                                    continue





                        if len(url_len)==9:
                            try:
                                driver.get(url)
                                pause(randint(5,7))
                                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'js-price-inner')))
                                driver.execute_script("window.scrollBy(0,5000)")
                                soup2 = BeautifulSoup(driver.page_source, 'lxml')

                                for x in soup2.find_all('div', class_='mainpage justify view-item-page'):
                                    product = dict()
                                    try:
                                        article = x.find('td', class_='property__table-value').text
                                        name =  x.find('h1', class_='product__title').text.split('\n')[1]
                                        if any(p['артикул'] == article or p['наименование'] == name for p in products):
                                            print('такое уже было, проверь ', url)
                                            continue
                                        else:
                                            product['артикул'] = article
                                            product['наименование'] = name
                                            
                                            try:
                                                product_price = x.find('span', class_='js-price-inner').text.split('\n')[1]
                                                product_price = product_price.replace(',', '.').replace('Р','').replace(' ','')
                                                if '.' not in product_price:
                                                    product_price += '.00'
                                                match = re.search(r'\d+\.\d+', product_price)
                                                if match:
                                                    product['цена'] = float(match.group(0))
                                                else:
                                                    product['цена'] = None
                                            except Exception:
                                                product['цена'] = None



                                            try:
                                                availability = x.find('div', class_='territory-choose__list-count').text.strip()
                                                if availability:
                                                    availability = availability.replace(',', '.').replace('шт.','').replace(' ','')
                                                    product['наличие'] = int(availability)
                                                else:
                                                    product['наличие'] = availability
                                            except Exception:
                                                product['наличие'] = availability



                                            products.append(product)
                                            print(product)
                                            print(f'{count}    КАРТОЧКА ОБРАБОТАНА')
                                            count += 1
                                        print('+++++++++++')
                                    except Exception as e:
                                        errors.extend((str(e), url))
                                        continue


                            except TimeoutException:
                                errors.extend('TimeoutExeption', url)
                                continue
                        else:
                            continue


                    except TimeoutException:
                        errors.extend('TimeoutExeption', url)
                        continue

                    except Exception as e:
                        errors.extend((str(e), url))
                        continue
        except Exception as e:
            errors.extend((str(e), url))
            print('ALARM!!!', url)
            continue

        finally:
            if errors:
                for error in errors:
                    print(f"Error: {error[0]}, URL: {error[1]}")
            else:
                print("No errors.")


for product in products:
    column_names.extend(list(set(product) - set(column_names)))


to_excel(products, column_names, file_name='ORG2, краткий сбор по наличию')
print('ФАЙЛ ЗАПИСАН')
print('-------------------------------------------------')
print('Обработка закончена:   ', str(datetime.now() - start_time).split('.')[0])
