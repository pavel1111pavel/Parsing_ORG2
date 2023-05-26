import os
import re
from datetime import datetime
from datetime import date
from random import randint
from time import sleep as pause
import csv
from typing import Any, Iterable, Mapping, Optional

from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, Side, PatternFill
from openpyxl.utils import get_column_letter
from selenium.webdriver import Chrome
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.wait import WebDriverWait
from tqdm import tqdm
from webdriver_manager.chrome import ChromeDriverManager
import openpyxl
import string

start_time = datetime.now()
total_pagen = []
with open('ссылки с пагинацией.txt', 'r') as file:

    urls = [x for x in file]
    x = [word.strip(string.punctuation) for word in str(urls).split()]
    for i in x:
        total_pagen.append(i)

def get_date_and_time() -> str:
    return datetime.now().strftime('%d.%m.%y %H-%M-%S')


def to_excel(data: Iterable[Mapping[str, Any]], column_names: Iterable[str],
             file_name: Optional[str] = "table") -> None:
    """
    Создаёт из итерируемого объекта и имён столбцов
    xlsx файл в папке "resulting files".
    """
    wb = Workbook()
    worksheet = wb.active
    side = Side(border_style='thin')
    border = Border(
        left=side,
        right=side,
        top=side,
        bottom=side
    )
    alignment = Alignment(
        horizontal='left',
        vertical='center'
    )
    column_widths = []

    for column, name in enumerate(column_names, 1):
        cell = worksheet.cell(
            column=column,
            row=1,
            value=name
        )
        cell.font = Font(name='Calibri', size=11, bold=True)
        cell.fill = PatternFill('solid', fgColor='C5D9F1')
        cell.border = border
        cell.alignment = alignment
        column_widths.append(len(name) + 2)

    for row, product in enumerate(data, 2):
        if not product:
            print(row)
            continue
        for column, name in enumerate(column_names, 1):
            cell = worksheet.cell(
                column=column,
                row=row,
                value=product.get(name, '')
            )
            cell.font = Font(name='Calibri', size=11, bold=False)
            cell.border = border
            cell.alignment = alignment
            column_widths[column -
                          1] = max(column_widths[column -
                                                 1], len(str(cell.value)))

    for i, column_width in enumerate(column_widths, 1):
        worksheet.column_dimensions[get_column_letter(
            i)].width = column_width + 1

    datetime_now = get_date_and_time()
    wb.save(f"resulting files/{file_name}____{datetime_now}.xlsx")


#начало работы парсера
options = webdriver.ChromeOptions()
options.page_load_strategy = 'none'

with webdriver.Chrome(options=options) as driver:
    driver.maximize_window()

    print("Найдено страниц в каталоге:", len(total_pagen))

    pause(1)

    products = []
    column_names = [
        'артикул',
        'наименование',
        'цена',
        'наличие'
    ]
    total_cards = []
    #обработка каталогов на покарточную пагинацию
    for url in total_pagen:
        driver.get(url)
        pause(4)

        try:
            window_handles = driver.window_handles
            # Переключаемся на второе окно
            driver.switch_to.window(window_handles[0])
            pause(2)
            element_present = EC.presence_of_element_located((By.CLASS_NAME, 'products__info-block'))
            WebDriverWait(driver, 15).until(element_present)
            soup = BeautifulSoup(driver.page_source, 'html.parser')
            cards_url = ['https://www.santech.ru' + x.find('a').get('href') for x in soup.find_all('div', class_='products__info-block')]
            print('try', cards_url, 'собранные таблички11111')
            total_cards.append(cards_url)
            pause(2)
            # element_present = EC.presence_of_element_located((By.CLASS_NAME, 'products__info-block'))
            # WebDriverWait(driver, 15).until(element_present)
        except TimeoutException:

            element_present = EC.presence_of_element_located((By.CLASS_NAME, 'products__info-block'))
            WebDriverWait(driver, 15).until(element_present)

            soup = BeautifulSoup(driver.page_source, 'html.parser')
            pause(2)
            cards_url = ['https://www.santech.ru' + x.find('a').get('href') for x in soup.find_all('div', class_='products__info-block')]
            print('except', cards_url, 'собранные таблички')
            total_cards.append(cards_url)
            pause(2)
            print('Timed out waiting for page to load')
            continue

    count = 1
    #работа с карточками, извлечение данных
    for card in total_cards:

        for url in card:


            try:
                driver.get(url)
                pause(randint(4,7))
                WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'js-price-inner')))
                driver.execute_script("window.scrollBy(0,5000)")
                soup1 = BeautifulSoup(driver.page_source, 'lxml')

                print(url)
                url_len = url.split('/')

                print(len(url_len))
                if len(url_len)==8:

                    for x in soup1.find_all('div', class_='var-header col-cont'):
                        product = dict()
                        try:
                            product['артикул'] = x.find('span', class_='variant-list__jde-code variant-list__jde-num').text

                        except Exception:
                            product['артикул'] = 'нет данных'

                        try:
                            product['наименование'] = x.find('span', class_='variant-list__info-title js-variant_title').text.split('\n')[1]
                        except Exception:
                            product['наименование'] = 'нет данных'

                        try:
                            product_price = x.find('span', class_='js-price-inner').text.split('\n')[1]
                            print(product_price)
                            product_price = product_price.replace(',', '.').replace('Р','').replace(' ','')
                            if '.' not in product_price:
                                product_price += '.00'
                            match = re.search(r'\d+\.\d+', product_price)
                            print(match)

                            if match:
                                product['цена'] = float(match.group(0))
                            else:
                                product['цена'] = None
                        except Exception:
                            product['цена'] = None

                        try:
                            product['наличие'] = x.find('a', class_='js-territories-popup').text
                        except Exception:
                            product['наличие'] = 'нет данных'

                        count += 1

                        products.append(product)
                        print(product)
                        print(f'{count}    КАРТОЧКА ОБРАБОТАНА')


                        print('+++++++++++')



                if len(url_len)==9:
                    try:
                        driver.get(url)
                        pause(randint(5,7))
                        WebDriverWait(driver, 30).until(EC.presence_of_element_located((By.CLASS_NAME, 'js-price-inner')))
                        driver.execute_script("window.scrollBy(0,5000)")
                        soup2 = BeautifulSoup(driver.page_source, 'lxml')

                        for x in soup2.find_all('div', class_='mainpage justify view-item-page'):
                            product = dict()
                            try:
                                product['артикул'] = x.find('td', class_='property__table-value').text

                            except Exception:
                                product['артикул'] = 'нет данных'

                            try:
                                product['наименование'] = x.find('h1', class_='product__title').text.split('\n')[1]

                            except Exception:
                                product['наименование'] = 'нет данных'



                            try:
                                product_price = x.find('span', class_='js-price-inner').text.split('\n')[1]
                                print(product_price)
                                product_price = product_price.replace(',', '.').replace('Р','').replace(' ','')
                                if '.' not in product_price:
                                    product_price += '.00'
                                match = re.search(r'\d+\.\d+', product_price)
                                print(match)

                                if match:
                                    product['цена'] = float(match.group(0))
                                else:
                                    product['цена'] = None
                            except Exception:
                                product['цена'] = None



                            try:
                                product['наличие'] = x.find('div', class_='territory-choose__list-count').text.split('\n').strip()
                                print(x.find('div', class_='territory-choose__list-count').text.split('\n').strip())
                            except Exception:
                                product['наличие'] = 'нет данных'



                            products.append(product)
                            print(product)
                            print(f'{count}    КАРТОЧКА ОБРАБОТАНА')
                            count += 1
                            print('---------------------')
                    except TimeoutException:
                        continue
                else:
                    continue


            except TimeoutException:
                continue

#не знаю зачем это
for product in products:
    column_names.extend(list(set(product) - set(column_names)))


to_excel(products, column_names, file_name='Видное')
print('ФАЙЛ ЗАПИСАН')
print('-------------------------------------------------')
print('Обработка закончена:   ', str(datetime.now() - start_time).split('.')[0])


